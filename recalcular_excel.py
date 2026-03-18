"""
recalcular_excel.py
Lee un Excel de evaluación ya generado, aplica normalizaciones sobre las predicciones
y recalcula métricas sin volver a llamar a la API.

Normalización aplicada:
  - pred "No especificado" → "No disponible" (case-insensitive)
"""

import re
import sys
import numpy as np
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Configuración ────────────────────────────────────────────────────────────
# Pasa el path del Excel como argumento o pon el nombre aquí:
if len(sys.argv) > 1:
    XLSX_INPUT = sys.argv[1]
else:
    XLSX_INPUT = "evaluacion_catalogacion_paralelo_20260314_115818.xlsx"

timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
XLSX_OUTPUT = f"evaluacion_recalculada_{timestamp}.xlsx"

TIPO_API_A_CSV = {
    "laptop":  "notebook",
    "aio":     "all in one (aio)",
    "desktop": "desktop",
    "otro":    "otro",
}

CAMPOS_NUMERICOS = {"nucleos", "hilos", "ram", "almacenamiento", "pantalla (pulgadas)"}


# ─── Funciones de normalización (idénticas al script principal) ───────────────

def normalizar(valor) -> str:
    if valor is None or (isinstance(valor, float) and np.isnan(valor)):
        return ""
    s = str(valor).strip().lower()
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"\b(gb|tb|mb|ghz|mhz)\s+\1\b", r"\1", s)
    return s


def normalizar_numero(valor) -> str:
    if valor is None or (isinstance(valor, float) and np.isnan(valor)):
        return ""
    match = re.search(r"(\d+(?:[.,]\d+)?)", str(valor))
    if match:
        num = match.group(1).replace(",", ".")
        try:
            return str(int(float(num)))
        except ValueError:
            return num
    return normalizar(valor)


def son_iguales(campo: str, pred: str, label: str) -> bool:
    campo_lower = campo.lower()
    if campo_lower == "tipo producto":
        pred_norm  = TIPO_API_A_CSV.get(normalizar(pred), normalizar(pred))
        label_norm = normalizar(label)
        return pred_norm == label_norm
    if campo_lower in CAMPOS_NUMERICOS:
        return normalizar_numero(pred) == normalizar_numero(label)
    return normalizar(pred) == normalizar(label)


def aplicar_normalizaciones(pred: str) -> str:
    """Normaliza variantes de 'no especificado' → 'no disponible'."""
    if normalizar(pred) == "no especificado":
        return "No disponible"
    return pred


# ─── Lectura ──────────────────────────────────────────────────────────────────

print(f"Leyendo: {XLSX_INPUT}")
df = pd.read_excel(XLSX_INPUT, sheet_name="Detalle")

# Columnas esperadas: #, Codigo Cotizacion, Proveedor, Tipo Real,
#                     Campo, Label, Prediccion, Correcto, Segundos API
print(f"  Filas en Detalle: {len(df)}")

# ─── Aplicar normalización ────────────────────────────────────────────────────

cambios = 0
nuevos_correcto = []
nuevas_pred = []

for _, row in df.iterrows():
    pred_orig  = str(row["Prediccion"]) if not pd.isna(row["Prediccion"]) else ""
    label_val  = str(row["Label"])      if not pd.isna(row["Label"])      else ""
    campo      = str(row["Campo"])

    pred_nuevo = aplicar_normalizaciones(pred_orig)
    if pred_nuevo != pred_orig:
        cambios += 1

    nuevas_pred.append(pred_nuevo)

    if not label_val:
        nuevos_correcto.append(None)
    else:
        nuevos_correcto.append(son_iguales(campo, pred_nuevo, label_val))

df["Prediccion"] = nuevas_pred
df["correcto_bool"] = nuevos_correcto
print(f"  Predicciones normalizadas: {cambios}")

# ─── Recalcular métricas ──────────────────────────────────────────────────────

campos = df["Campo"].unique().tolist()

metricas = []
for campo in campos:
    sub = df[(df["Campo"] == campo) & (df["correcto_bool"].notna())]
    if len(sub) == 0:
        metricas.append({"campo": campo, "accuracy": None, "correctas": 0, "evaluadas": 0,
                         "sin_label": len(df[df["Campo"] == campo])})
        continue
    correctas = sub["correcto_bool"].sum()
    total = len(sub)
    metricas.append({
        "campo":     campo,
        "accuracy":  round(correctas / total, 4),
        "correctas": int(correctas),
        "evaluadas": total,
        "sin_label": len(df[df["Campo"] == campo]) - total,
    })

accs = [m["accuracy"] for m in metricas if m["accuracy"] is not None]
acc_general = round(sum(accs) / len(accs), 4) if accs else None

print(f"\nAccuracy general: {acc_general:.2%}" if acc_general else "\nAccuracy general: N/A")
for m in metricas:
    acc_str = f"{m['accuracy']:.2%}" if m["accuracy"] is not None else "N/A"
    print(f"  {m['campo']:<30} {acc_str}  ({m['correctas']}/{m['evaluadas']})")

# ─── Generar Excel ────────────────────────────────────────────────────────────

wb = Workbook()

verde       = PatternFill("solid", fgColor="C6EFCE")
rojo        = PatternFill("solid", fgColor="FFC7CE")
gris        = PatternFill("solid", fgColor="D9D9D9")
azul_hdr    = PatternFill("solid", fgColor="1F4E79")
amarillo    = PatternFill("solid", fgColor="FFEB9C")
font_hdr    = Font(color="FFFFFF", bold=True)
font_bold   = Font(bold=True)
centro      = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

def estilo_hdr(cell):
    cell.fill = azul_hdr; cell.font = font_hdr
    cell.alignment = centro; cell.border = thin_border

def estilo_celda(cell, fill=None):
    if fill: cell.fill = fill
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = thin_border

# Hoja 1: Detalle
ws = wb.active
ws.title = "Detalle"
cols_det   = ["#", "Codigo Cotizacion", "Proveedor", "Tipo Real",
              "Campo", "Label", "Prediccion", "Correcto", "Segundos API"]
for ci, h in enumerate(cols_det, 1):
    estilo_hdr(ws.cell(row=1, column=ci, value=h))

orig_cols = ["#", "Codigo Cotizacion", "Proveedor", "Tipo Real",
             "Campo", "Label", "Prediccion", "Correcto", "Segundos API"]

for ri, (_, row) in enumerate(df.iterrows(), 2):
    correcto = row["correcto_bool"]
    fill = verde if correcto is True else (rojo if correcto is False else gris)
    vals = [
        row.get("#", row.get("Unnamed: 0", ri - 1)),
        row.get("Codigo Cotizacion", ""),
        row.get("Proveedor", ""),
        row.get("Tipo Real", ""),
        row["Campo"],
        row["Label"],
        row["Prediccion"],
        "SI" if correcto is True else ("NO" if correcto is False else "S/L"),
        row.get("Segundos API", ""),
    ]
    for ci, val in enumerate(vals, 1):
        cell = ws.cell(row=ri, column=ci, value=val)
        if ci in (6, 7, 8):  # Label, Prediccion, Correcto
            estilo_celda(cell, fill=fill)
        else:
            estilo_celda(cell)

for i, w in enumerate([6, 20, 30, 15, 28, 25, 35, 10, 12], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"

# Hoja 2: Metricas
ws_met = wb.create_sheet("Metricas")
ws_met.merge_cells("A1:F1")
c = ws_met.cell(row=1, column=1, value="Evaluación Recalculada — Normalización 'No especificado' → 'No disponible'")
c.font = Font(color="FFFFFF", bold=True, size=12); c.alignment = centro; c.fill = azul_hdr

ws_met.cell(row=3, column=1, value="Accuracy General (promedio campos)").font = font_bold
ws_met.cell(row=3, column=2, value=f"{acc_general:.2%}" if acc_general else "N/A").fill = amarillo
ws_met.cell(row=4, column=1, value="Fuente Excel").font = font_bold
ws_met.cell(row=4, column=2, value=XLSX_INPUT)
ws_met.cell(row=5, column=1, value="Timestamp recálculo").font = font_bold
ws_met.cell(row=5, column=2, value=timestamp)
ws_met.cell(row=6, column=1, value="Cambios de prediccion").font = font_bold
ws_met.cell(row=6, column=2, value=cambios)

for ci, h in enumerate(["Campo", "Accuracy", "Correctas", "Evaluadas", "Sin Label"], 1):
    estilo_hdr(ws_met.cell(row=8, column=ci, value=h))

for ri, m in enumerate(metricas, 9):
    acc = m["accuracy"]
    ws_met.cell(row=ri, column=1, value=m["campo"]).font = font_bold
    c2 = ws_met.cell(row=ri, column=2, value=f"{acc:.2%}" if acc else "N/A")
    if acc is not None:
        c2.fill = verde if acc >= 0.8 else (amarillo if acc >= 0.5 else rojo)
    ws_met.cell(row=ri, column=3, value=m["correctas"])
    ws_met.cell(row=ri, column=4, value=m["evaluadas"])
    ws_met.cell(row=ri, column=5, value=m["sin_label"])

for i, w in enumerate([30, 12, 12, 12, 12], 1):
    ws_met.column_dimensions[get_column_letter(i)].width = w

wb.save(XLSX_OUTPUT)
print(f"\nExcel guardado: {XLSX_OUTPUT}")
