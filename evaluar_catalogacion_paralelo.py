"""
evaluar_catalogacion_paralelo.py
Evaluación de precisión del pipeline /catalogar (Compra Ágil) con llamadas en paralelo.

Llama al endpoint con muestras del CSV etiquetado en lotes de N_PARALELO,
compara predicciones vs. labels reales y genera un Excel con:
  - Hoja "Detalle"     : una fila por (muestra × campo) con label, pred y estado
  - Hoja "Por_Muestra" : una fila por muestra, columnas agrupadas por campo
  - Hoja "Metricas"    : accuracy por campo + accuracy general
"""

import re
import json
import time
import logging
import concurrent.futures
from datetime import datetime

import numpy as np
import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuración
# ---------------------------------------------------------------------------
API_URL   = "https://aichilecompra-production.up.railway.app"
API_KEY   = "1234"
N_PARALELO = 10   # llamadas simultáneas por lote

CSV_INPUT      = "sample_OC_100_rows_computadores.csv"
TIPOS_PRODUCTO = ["Notebook", "Desktop", "All in One (AIO)"]
N_MUESTRAS     = 100

LABEL_FIELDS = [
    "Modalidad",
    "Tipo Producto",
    "Marca",
    "Procesador",
    "Nucleos",
    "Hilos",
    "RAM (GB)",
    "Tipo RAM",
    "Sistema Operativo",
    "Tipo Almacenamiento",
    "Almacenamiento (GB)",
    "Pantalla (Pulgadas)",
]

CAMPOS_MANUALES_API = [f for f in LABEL_FIELDS]

TIPO_API_A_CSV = {
    "laptop":  "notebook",
    "aio":     "all in one (aio)",
    "desktop": "desktop",
    "otro":    "otro",
}

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
log_file    = f"evaluar_catalogacion_paralelo_{timestamp}.log"
xlsx_output = f"evaluacion_catalogacion_paralelo_{timestamp}.xlsx"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(log_file, encoding="utf-8"),
    ],
)
logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------

def limpiar_texto(valor) -> str:
    if pd.isna(valor):
        return ""
    return str(valor).strip()


def normalizar(valor) -> str:
    if valor is None or (isinstance(valor, float) and np.isnan(valor)):
        return ""
    s = str(valor).strip().lower()
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"\b(gb|tb|mb|ghz|mhz)\s+\1\b", r"\1", s)
    return s


def normalizar_numero(valor) -> str:
    if valor is None or (isinstance(valor, float) and np.isnan(valor)):
        return ""
    match = re.search(r"(\d+(?:[.,]\d+)?)", str(valor))
    if match:
        num = match.group(1).replace(",", ".")
        try:
            return str(int(float(num)))
        except ValueError:
            return num
    return normalizar(valor)


CAMPOS_NUMERICOS = {"nucleos", "hilos", "ram (gb)", "almacenamiento (gb)", "pantalla (pulgadas)"}


def son_iguales(campo: str, pred: str, label: str) -> bool:
    campo_lower = campo.lower()
    if campo_lower == "tipo producto":
        pred_norm  = TIPO_API_A_CSV.get(normalizar(pred), normalizar(pred))
        label_norm = normalizar(label)
        return pred_norm == label_norm
    if campo_lower in CAMPOS_NUMERICOS:
        return normalizar_numero(pred) == normalizar_numero(label)
    return normalizar(pred) == normalizar(label)


# ---------------------------------------------------------------------------
# Llamada al API
# ---------------------------------------------------------------------------

def llamar_api(iteracion: int, row: pd.Series) -> dict:
    """Llama a /catalogar y retorna un dict con todos los datos necesarios para evaluar."""
    codigo_cot    = limpiar_texto(row["CodigoExterno"])
    rut_proveedor = limpiar_texto(row["RutSucursal"])
    nombre_prov   = limpiar_texto(row["NombreProveedor"])
    tipo_real     = limpiar_texto(row["Tipo Producto"])

    payload_dict = {
        "Categoria": "Computadores",
        "DescripcionProductoComprador": limpiar_texto(row["EspecificacionComprador"]),
        "DescripcionProductoProveedor": limpiar_texto(row["EspecificacionProveedor"]),
        "productoname": limpiar_texto(row["NombreroductoGenerico"]),
    }

    body = {
        "payload": payload_dict,
        "codigo_cotizacion": codigo_cot,
        "rut_proveedor": rut_proveedor,
        "use_diccionarios": True,
        "campos_manuales": CAMPOS_MANUALES_API,
    }

    t_ini = time.time()
    try:
        resp = requests.post(
            f"{API_URL}/catalogar",
            headers={"x-api-key": API_KEY, "Content-Type": "application/json"},
            json=body,
            timeout=360,
        )
        resp.raise_for_status()
        data = resp.json()
    except Exception as exc:
        t_seg = round(time.time() - t_ini, 2)
        logger.error(f"  [{iteracion:02d}] Error API {codigo_cot}: {exc}")
        return {
            "iteracion": iteracion,
            "codigo_cot": codigo_cot,
            "rut_proveedor": rut_proveedor,
            "nombre_prov": nombre_prov,
            "tipo_real": tipo_real,
            "row": row,
            "data": {"error": str(exc)},
            "segundos": t_seg,
        }

    t_seg = round(time.time() - t_ini, 2)
    logger.info(f"  [{iteracion:02d}/{N_MUESTRAS}] ✓ {tipo_real} | {codigo_cot} | {t_seg}s | Tipo: {data.get('resultado', {}).get('Tipo', 'N/A')}")
    return {
        "iteracion": iteracion,
        "codigo_cot": codigo_cot,
        "rut_proveedor": rut_proveedor,
        "nombre_prov": nombre_prov,
        "tipo_real": tipo_real,
        "row": row,
        "data": data,
        "segundos": t_seg,
    }


def procesar_resultado(res: dict) -> tuple[list[dict], dict]:
    """Convierte el resultado de una llamada en filas de detalle y fila wide."""
    iteracion  = res["iteracion"]
    codigo_cot = res["codigo_cot"]
    nombre_prov = res["nombre_prov"]
    tipo_real  = res["tipo_real"]
    t_seg      = res["segundos"]
    row        = res["row"]
    data       = res["data"]

    fila_wide = {
        "iteracion": iteracion,
        "codigo_cotizacion": codigo_cot,
        "proveedor": nombre_prov,
        "tipo_real": tipo_real,
        "segundos": t_seg,
        "exito_api": False,
    }

    if "error" in data:
        logger.error(f"  [{iteracion:02d}] Muestra fallida: {data['error']}")
        return [], fila_wide

    resultado = data.get("resultado") or {}
    fila_wide["exito_api"] = data.get("success", False)

    filas_detalle = []
    for campo in LABEL_FIELDS:
        label_val = limpiar_texto(row.get(campo, ""))
        pred_val  = limpiar_texto(resultado.get("Tipo", "") if campo == "Tipo Producto" else resultado.get(campo, ""))
        correcto  = son_iguales(campo, pred_val, label_val) if label_val else None

        estado_str = "OK" if correcto else ("--" if correcto is None else "FAIL")
        logger.info(f"      [{estado_str}] {campo:<25} | label={label_val!r:25} | pred={pred_val!r}")

        filas_detalle.append({
            "iteracion":         iteracion,
            "codigo_cotizacion": codigo_cot,
            "proveedor":         nombre_prov,
            "tipo_real":         tipo_real,
            "campo":             campo,
            "label":             label_val,
            "prediccion":        pred_val,
            "correcto":          correcto,
            "segundos_api":      t_seg,
        })

        fila_wide[f"{campo}_label"] = label_val
        fila_wide[f"{campo}_pred"]  = pred_val
        fila_wide[f"{campo}_ok"]    = correcto

    return filas_detalle, fila_wide


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    logger.info("=" * 70)
    logger.info(f"EVALUACIÓN /catalogar — Computadores (paralelo N={N_PARALELO})")
    logger.info(f"Salida: {xlsx_output}")
    logger.info("=" * 70)

    # ------------------------------------------------------------------
    # 1. Cargar CSV y muestrear
    # ------------------------------------------------------------------
    df = pd.read_csv(CSV_INPUT)
    df_filtrado = df[df["Tipo Producto"].isin(TIPOS_PRODUCTO)].copy()
    df_filtrado = df_filtrado.dropna(subset=["CodigoExterno", "RutSucursal"])

    muestras = (
        df_filtrado
        .groupby("Tipo Producto", group_keys=False)
        .apply(lambda g: g.sample(min(len(g), N_MUESTRAS), random_state=42))
        .sample(frac=1, random_state=42)
        .head(N_MUESTRAS)
        .reset_index(drop=True)
    )

    logger.info(f"Muestras seleccionadas: {len(muestras)}")
    logger.info(f"Distribución:\n{muestras['Tipo Producto'].value_counts().to_string()}\n")

    # ------------------------------------------------------------------
    # 2. Evaluar en lotes paralelos
    # ------------------------------------------------------------------
    filas_detalle: list[dict] = []
    filas_muestras: list[dict] = []

    t_global = time.time()
    n_total  = len(muestras)
    n_lotes  = (n_total + N_PARALELO - 1) // N_PARALELO

    for lote_idx in range(n_lotes):
        inicio = lote_idx * N_PARALELO
        fin    = min(inicio + N_PARALELO, n_total)
        lote   = [(inicio + i, muestras.iloc[inicio + i]) for i in range(fin - inicio)]

        logger.info("-" * 70)
        logger.info(f"  Lote {lote_idx + 1}/{n_lotes} — muestras {inicio + 1}-{fin}")

        t_lote = time.time()
        with concurrent.futures.ThreadPoolExecutor(max_workers=N_PARALELO) as executor:
            futures = {executor.submit(llamar_api, idx + 1, row): idx for idx, row in lote}
            resultados_lote = []
            for future in concurrent.futures.as_completed(futures):
                resultados_lote.append(future.result())

        # Ordenar por iteracion para logging consistente
        resultados_lote.sort(key=lambda r: r["iteracion"])
        logger.info(f"  Lote completado en {round(time.time() - t_lote, 1)}s")

        for res in resultados_lote:
            det, wide = procesar_resultado(res)
            filas_detalle.extend(det)
            filas_muestras.append(wide)

    t_total = round(time.time() - t_global, 1)
    logger.info(f"\n  Tiempo total: {t_total}s  |  promedio por muestra: {round(t_total / n_total, 1)}s")

    # ------------------------------------------------------------------
    # 3. Calcular métricas
    # ------------------------------------------------------------------
    df_det = pd.DataFrame(filas_detalle)

    metricas: list[dict] = []
    for campo in LABEL_FIELDS:
        sub = df_det[(df_det["campo"] == campo) & (df_det["correcto"].notna())]
        if len(sub) == 0:
            metricas.append({
                "campo": campo, "accuracy": None,
                "correctas": 0, "evaluadas": 0,
                "sin_label": len(df_det[df_det["campo"] == campo]),
            })
            continue
        correctas = sub["correcto"].sum()
        total     = len(sub)
        metricas.append({
            "campo":     campo,
            "accuracy":  round(correctas / total, 4),
            "correctas": int(correctas),
            "evaluadas": total,
            "sin_label": len(df_det[df_det["campo"] == campo]) - total,
        })

    accs_validas = [m["accuracy"] for m in metricas if m["accuracy"] is not None]
    acc_general  = round(sum(accs_validas) / len(accs_validas), 4) if accs_validas else None

    df_det_eval = df_det[df_det["correcto"].notna()]
    acc_por_muestra = (
        df_det_eval.groupby("iteracion")["correcto"].mean().mean()
        if not df_det_eval.empty else None
    )

    logger.info("\n" + "=" * 70)
    logger.info("MÉTRICAS FINALES")
    logger.info(f"  Accuracy general (promedio campos): {acc_general:.2%}" if acc_general else "  N/A")
    for m in metricas:
        acc_str = f"{m['accuracy']:.2%}" if m["accuracy"] is not None else "N/A"
        logger.info(f"  {m['campo']:<30} {acc_str} ({m['correctas']}/{m['evaluadas']})")

    # ------------------------------------------------------------------
    # 4. Generar Excel (idéntico al original)
    # ------------------------------------------------------------------
    wb = Workbook()

    verde       = PatternFill("solid", fgColor="C6EFCE")
    rojo        = PatternFill("solid", fgColor="FFC7CE")
    gris        = PatternFill("solid", fgColor="D9D9D9")
    azul_hdr    = PatternFill("solid", fgColor="1F4E79")
    amarillo    = PatternFill("solid", fgColor="FFEB9C")
    font_hdr    = Font(color="FFFFFF", bold=True)
    font_bold   = Font(bold=True)
    centro      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    def estilo_hdr(cell):
        cell.fill      = azul_hdr
        cell.font      = font_hdr
        cell.alignment = centro
        cell.border    = thin_border

    def estilo_celda(cell, fill=None):
        if fill:
            cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border    = thin_border

    # ── Hoja 1: Detalle ──────────────────────────────────────────────
    ws_det = wb.active
    ws_det.title = "Detalle"

    cols_det   = ["iteracion", "codigo_cotizacion", "proveedor", "tipo_real",
                  "campo", "label", "prediccion", "correcto", "segundos_api"]
    headers_det = ["#", "Codigo Cotizacion", "Proveedor", "Tipo Real",
                   "Campo", "Label", "Prediccion", "Correcto", "Segundos API"]

    for col_idx, hdr in enumerate(headers_det, 1):
        estilo_hdr(ws_det.cell(row=1, column=col_idx, value=hdr))

    for row_idx, fila in enumerate(filas_detalle, 2):
        correcto = fila.get("correcto")
        fill = verde if correcto is True else (rojo if correcto is False else gris)
        for col_idx, col in enumerate(cols_det, 1):
            val = fila.get(col, "")
            if col == "correcto":
                val = "SI" if correcto is True else ("NO" if correcto is False else "S/L")
            cell = ws_det.cell(row=row_idx, column=col_idx, value=val)
            estilo_celda(cell, fill=fill if col in ("label", "prediccion", "correcto") else None)

    for i, w in enumerate([6, 20, 30, 15, 28, 25, 35, 10, 12], 1):
        ws_det.column_dimensions[get_column_letter(i)].width = w
    ws_det.freeze_panes = "A2"

    # ── Hoja 2: Por_Muestra ──────────────────────────────────────────
    ws_wide = wb.create_sheet("Por_Muestra")
    base_cols = ["iteracion", "codigo_cotizacion", "proveedor", "tipo_real", "segundos", "exito_api"]
    base_hdrs = ["#", "Codigo Cotizacion", "Proveedor", "Tipo Real", "Segundos", "Exito API"]

    campo_cols, campo_hdrs = [], []
    for campo in LABEL_FIELDS:
        campo_cols += [f"{campo}_label", f"{campo}_pred", f"{campo}_ok"]
        campo_hdrs += [f"{campo} [Label]", f"{campo} [Pred]", f"{campo} [OK]"]

    all_cols_wide = base_cols + campo_cols
    all_hdrs_wide = base_hdrs + campo_hdrs

    for col_idx, hdr in enumerate(all_hdrs_wide, 1):
        estilo_hdr(ws_wide.cell(row=1, column=col_idx, value=hdr))

    for row_idx, fila in enumerate(filas_muestras, 2):
        for col_idx, col in enumerate(all_cols_wide, 1):
            val = fila.get(col, "")
            if col.endswith("_ok"):
                fill = verde if val is True else (rojo if val is False else gris)
                val  = "SI" if val is True else ("NO" if val is False else "S/L")
                estilo_celda(ws_wide.cell(row=row_idx, column=col_idx, value=val), fill=fill)
            else:
                estilo_celda(ws_wide.cell(row=row_idx, column=col_idx, value=val))

    for col_idx in range(1, len(all_cols_wide) + 1):
        hdr = all_hdrs_wide[col_idx - 1]
        w   = 8 if "[OK]" in hdr else (22 if "[Label]" in hdr or "[Pred]" in hdr else 16)
        ws_wide.column_dimensions[get_column_letter(col_idx)].width = w
    ws_wide.freeze_panes = "G2"

    # ── Hoja 3: Metricas ─────────────────────────────────────────────
    ws_met = wb.create_sheet("Metricas")
    ws_met.merge_cells("A1:F1")
    cell_titulo = ws_met.cell(row=1, column=1,
                               value="Evaluación de Precisión — Catalogación de Computadores (Paralelo)")
    cell_titulo.font      = Font(color="FFFFFF", bold=True, size=13)
    cell_titulo.alignment = centro
    cell_titulo.fill      = azul_hdr

    ws_met.cell(row=3, column=1, value="Accuracy General (promedio campos)").font = font_bold
    ws_met.cell(row=3, column=2, value=f"{acc_general:.2%}" if acc_general else "N/A").fill = amarillo

    ws_met.cell(row=4, column=1, value="Muestras evaluadas").font = font_bold
    ws_met.cell(row=4, column=2, value=len(muestras))

    ws_met.cell(row=5, column=1, value="Paralelo (N workers)").font = font_bold
    ws_met.cell(row=5, column=2, value=N_PARALELO)

    ws_met.cell(row=6, column=1, value="Tiempo total (s)").font = font_bold
    ws_met.cell(row=6, column=2, value=t_total)

    ws_met.cell(row=7, column=1, value="Timestamp").font = font_bold
    ws_met.cell(row=7, column=2, value=timestamp)

    hdrs_met = ["Campo", "Accuracy", "Correctas", "Evaluadas", "Sin Label", "Barra"]
    for col_idx, hdr in enumerate(hdrs_met, 1):
        estilo_hdr(ws_met.cell(row=9, column=col_idx, value=hdr))

    for row_idx, m in enumerate(metricas, 10):
        acc = m["accuracy"]
        ws_met.cell(row=row_idx, column=1, value=m["campo"]).font = font_bold

        acc_cell = ws_met.cell(row=row_idx, column=2,
                               value=f"{acc:.2%}" if acc is not None else "N/A")
        acc_cell.fill      = (verde if acc >= 0.8 else (amarillo if acc >= 0.5 else rojo)) if acc is not None else gris
        acc_cell.alignment = centro

        ws_met.cell(row=row_idx, column=3, value=m["correctas"]).alignment = centro
        ws_met.cell(row=row_idx, column=4, value=m["evaluadas"]).alignment = centro
        ws_met.cell(row=row_idx, column=5, value=m["sin_label"]).alignment = centro

        if acc is not None:
            n_llenos = int(round(acc * 20))
            ws_met.cell(row=row_idx, column=6,
                        value="█" * n_llenos + "░" * (20 - n_llenos)).font = Font(name="Consolas")

    fila_total = 10 + len(metricas)
    ws_met.cell(row=fila_total, column=1, value="TOTAL / PROMEDIO").font = Font(bold=True, size=11)
    cell_ta = ws_met.cell(row=fila_total, column=2,
                          value=f"{acc_general:.2%}" if acc_general else "N/A")
    cell_ta.font = Font(bold=True)
    cell_ta.fill = amarillo
    cell_ta.alignment = centro
    ws_met.cell(row=fila_total, column=3, value=sum(m["correctas"] for m in metricas)).font = Font(bold=True)
    ws_met.cell(row=fila_total, column=4, value=sum(m["evaluadas"] for m in metricas)).font = Font(bold=True)

    for col, w in zip(["A","B","C","D","E","F"], [32, 14, 12, 12, 12, 26]):
        ws_met.column_dimensions[col].width = w

    # ------------------------------------------------------------------
    # 5. Guardar
    # ------------------------------------------------------------------
    wb.save(xlsx_output)
    logger.info(f"\nExcel guardado: {xlsx_output}")
    logger.info(f"Log guardado:   {log_file}")
    logger.info("=" * 70)
    logger.info("FIN EVALUACIÓN")


if __name__ == "__main__":
    main()
