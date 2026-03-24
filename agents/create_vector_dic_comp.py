import os
import sys
import logging
from typing import List, Dict
from docx import Document
from openpyxl import load_workbook

# Añadir el directorio raíz al path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from agents.get_vectorstore import VectorStoreManager

# Configuración de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Rutas
DICCIONARIOS_PATH = "diccionarios/Computadores"
CACHE_PATH = "cache/Computadores"
DOCX_FILE = "diccionario_features.docx"
XLSX_FILE = "diccionario_procesador.xlsx"


def extract_text_from_docx(docx_path: str) -> str:
    """
    Extrae todo el texto de un archivo DOCX como un solo chunk.

    Args:
        docx_path (str): Ruta al archivo DOCX

    Returns:
        str: Texto completo del documento
    """
    try:
        doc = Document(docx_path)
        text = '\n'.join([paragraph.text for paragraph in doc.paragraphs if paragraph.text.strip()])
        logging.info(f"Extracted text from DOCX: {len(text)} characters")
        return text
    except Exception as e:
        logging.error(f"Error extracting text from DOCX {docx_path}: {e}")
        raise


def extract_data_from_xlsx(xlsx_path: str) -> List[Dict[str, str]]:
    """
    Extrae datos de un archivo XLSX, creando un documento por cada fila.
    Cada documento contiene el nombre de la columna y el valor de la celda.

    Args:
        xlsx_path (str): Ruta al archivo XLSX

    Returns:
        List[Dict]: Lista de diccionarios con 'text' y 'metadata' para cada fila
    """
    try:
        workbook = load_workbook(xlsx_path, data_only=True)
        sheet = workbook.active

        # Obtener headers (primera fila)
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value))

        if not headers:
            raise ValueError("No se encontraron headers en el archivo XLSX")

        logging.info(f"Found {len(headers)} columns: {headers}")

        # Procesar cada fila (empezando desde la segunda)
        documents = []
        row_count = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Crear un documento por cada celda con valor
            for col_idx, cell_value in enumerate(row):
                if col_idx < len(headers) and cell_value is not None:
                    column_name = headers[col_idx]

                    # Crear texto descriptivo
                    text = f"{column_name}: {cell_value}"

                    documents.append({
                        'text': text,
                        'column': column_name,
                        'value': str(cell_value),
                        'row_number': row_count + 2  # +2 porque empezamos en fila 2
                    })

            row_count += 1

        logging.info(f"Extracted {len(documents)} documents from {row_count} rows")
        return documents

    except Exception as e:
        logging.error(f"Error extracting data from XLSX {xlsx_path}: {e}")
        raise


def create_computadores_vectorstore(
    diccionarios_path: str = DICCIONARIOS_PATH,
    cache_path: str = CACHE_PATH,
    collection_name: str = "diccionario_computadores"
) -> VectorStoreManager:
    """
    Crea un vector store con los diccionarios de computadores.

    Args:
        diccionarios_path (str): Ruta a la carpeta de diccionarios
        cache_path (str): Ruta donde guardar el vector store
        collection_name (str): Nombre de la colección

    Returns:
        VectorStoreManager: Gestor del vector store creado

    Example:
        >>> manager = create_computadores_vectorstore()
        >>> results = manager.similarity_search("procesador intel core i5")
    """
    # Crear directorio de cache si no existe
    os.makedirs(cache_path, exist_ok=True)

    # Rutas completas a los archivos
    docx_path = os.path.join(diccionarios_path, DOCX_FILE)
    xlsx_path = os.path.join(diccionarios_path, XLSX_FILE)

    # Verificar que los archivos existan
    if not os.path.exists(docx_path):
        raise FileNotFoundError(f"No se encontró el archivo DOCX: {docx_path}")
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"No se encontró el archivo XLSX: {xlsx_path}")

    logging.info("=== Iniciando creación de vector store ===")

    # Inicializar vector store manager
    manager = VectorStoreManager(
        collection_name=collection_name,
        persist_directory=cache_path
    )

    # Lista para almacenar todos los textos y metadatas
    all_texts = []
    all_metadatas = []

    # 1. Procesar archivo DOCX (Features)
    logging.info(f"Procesando archivo DOCX: {DOCX_FILE}")
    docx_text = extract_text_from_docx(docx_path)

    all_texts.append(docx_text)
    all_metadatas.append({
        "origen": "Features",
        "tipo": "diccionario",
        "archivo": DOCX_FILE,
        "formato": "docx"
    })

    # 2. Procesar archivo XLSX (Procesador)
    logging.info(f"Procesando archivo XLSX: {XLSX_FILE}")
    xlsx_documents = extract_data_from_xlsx(xlsx_path)

    for doc in xlsx_documents:
        all_texts.append(doc['text'])
        all_metadatas.append({
            "origen": "Procesador",
            "tipo": "diccionario",
            "archivo": XLSX_FILE,
            "formato": "xlsx",
            "column": doc['column'],
            "row_number": doc['row_number']
        })

    # 3. Añadir todos los documentos al vector store en lotes
    logging.info(f"Añadiendo {len(all_texts)} documentos al vector store...")

    # ChromaDB tiene un límite de batch size, procesamos en lotes
    batch_size = 5000
    doc_ids = []

    for i in range(0, len(all_texts), batch_size):
        batch_texts = all_texts[i:i + batch_size]
        batch_metadatas = all_metadatas[i:i + batch_size]

        logging.info(f"Procesando lote {i//batch_size + 1}/{(len(all_texts) + batch_size - 1)//batch_size}...")
        batch_ids = manager.add_documents(batch_texts, batch_metadatas)
        doc_ids.extend(batch_ids)

    logging.info(f"✓ Vector store creado exitosamente")
    logging.info(f"  - Total documentos: {len(doc_ids)}")
    logging.info(f"  - Documentos Features (DOCX): 1")
    logging.info(f"  - Documentos Procesador (XLSX): {len(xlsx_documents)}")
    logging.info(f"  - Ubicación: {cache_path}")
    logging.info(f"  - Colección: {collection_name}")

    return manager


def load_computadores_vectorstore(
    cache_path: str = CACHE_PATH,
    collection_name: str = "diccionario_computadores"
) -> VectorStoreManager:
    """
    Carga el vector store de diccionarios de computadores.

    Args:
        cache_path (str): Ruta donde está guardado el vector store
        collection_name (str): Nombre de la colección

    Returns:
        VectorStoreManager: Gestor del vector store cargado

    Example:
        >>> manager = load_computadores_vectorstore()
        >>> results = manager.similarity_search("características de RAM")
    """
    if not os.path.exists(cache_path):
        raise FileNotFoundError(
            f"No se encontró el vector store en {cache_path}. "
            f"Ejecuta create_computadores_vectorstore() primero."
        )

    manager = VectorStoreManager(
        collection_name=collection_name,
        persist_directory=cache_path
    )

    logging.info(f"✓ Vector store cargado desde {cache_path}")
    logging.info(f"  - Total documentos: {manager.get_collection_count()}")

    return manager


def search_diccionario(query: str,
                       k: int = 5,
                       filter_origen: str = None,
                       cache_path: str = CACHE_PATH,
                       collection_name: str = "diccionario_computadores"):
    """
    Función helper para buscar en el diccionario de computadores.

    Args:
        query (str): Consulta de búsqueda
        k (int): Número de resultados
        filter_origen (str, optional): Filtrar por origen ('Features' o 'Procesador')
        cache_path (str): Ruta al cache
        collection_name (str): Nombre de la colección

    Returns:
        List[Document]: Resultados de la búsqueda

    Example:
        >>> results = search_diccionario("procesador intel", k=3, filter_origen="Procesador")
        >>> for doc in results:
        >>>     print(doc.page_content)
    """
    manager = load_computadores_vectorstore(cache_path, collection_name)

    filter_dict = None
    if filter_origen:
        filter_dict = {"origen": filter_origen}

    results = manager.similarity_search(query, k=k, filter=filter_dict)
    return results


# Script principal
if __name__ == "__main__":
    import sys

    # Determinar acción
    action = sys.argv[1] if len(sys.argv) > 1 else "create"

    if action == "create":
        print("\n=== Creando vector store de diccionarios de computadores ===\n")

        try:
            manager = create_computadores_vectorstore()

            print("\n=== Vector store creado exitosamente ===")
            print(f"Total documentos en colección: {manager.get_collection_count()}")

            # Mostrar ejemplo de búsqueda
            print("\n=== Ejemplo de búsqueda ===")
            print("Query: 'procesador intel core'")
            results = manager.similarity_search("procesador intel core", k=3)

            for i, doc in enumerate(results, 1):
                print(f"\nResultado {i}:")
                print(f"  Origen: {doc.metadata.get('origen', 'N/A')}")
                print(f"  Contenido: {doc.page_content[:150]}...")

        except Exception as e:
            logging.error(f"Error creando vector store: {e}")
            sys.exit(1)

    elif action == "search":
        if len(sys.argv) < 3:
            print("Usage: python create_vector_dic_comp.py search <query> [k] [origen]")
            print("\nExamples:")
            print("  python create_vector_dic_comp.py search 'procesador intel' 5")
            print("  python create_vector_dic_comp.py search 'RAM memoria' 3 Procesador")
            sys.exit(1)

        query = sys.argv[2]
        k = int(sys.argv[3]) if len(sys.argv) > 3 else 5
        filter_origen = sys.argv[4] if len(sys.argv) > 4 else None

        print(f"\n=== Búsqueda en diccionario ===")
        print(f"Query: {query}")
        print(f"K: {k}")
        if filter_origen:
            print(f"Filtro origen: {filter_origen}")

        try:
            results = search_diccionario(query, k=k, filter_origen=filter_origen)

            print(f"\n=== Resultados ({len(results)}) ===")
            for i, doc in enumerate(results, 1):
                print(f"\nResultado {i}:")
                print(f"  Origen: {doc.metadata.get('origen', 'N/A')}")
                print(f"  Archivo: {doc.metadata.get('archivo', 'N/A')}")
                if 'column' in doc.metadata:
                    print(f"  Columna: {doc.metadata['column']}")
                print(f"  Contenido: {doc.page_content}")

        except Exception as e:
            logging.error(f"Error buscando: {e}")
            sys.exit(1)

    elif action == "info":
        try:
            manager = load_computadores_vectorstore()
            print(f"\n=== Información del vector store ===")
            print(f"Total documentos: {manager.get_collection_count()}")
            print(f"Ubicación: {CACHE_PATH}")
            print(f"Colección: diccionario_computadores")

            # Contar por origen
            results_features = manager.similarity_search(".", k=1000, filter={"origen": "Features"})
            results_procesador = manager.similarity_search(".", k=1000, filter={"origen": "Procesador"})

            print(f"\nDocumentos por origen:")
            print(f"  - Features: {len(results_features)}")
            print(f"  - Procesador: {len(results_procesador)}")

        except Exception as e:
            logging.error(f"Error obteniendo información: {e}")
            sys.exit(1)

    else:
        print("Acción no reconocida. Usa: create, search, o info")
        sys.exit(1)
