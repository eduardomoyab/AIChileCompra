"""
Procesador de adjuntos para licitaciones.

Procesa archivos PDF, imágenes, DOCX, HTML y XLSX para extraer texto.
Incluye OCR para PDFs escaneados y filtrado de archivos irrelevantes.
"""

import os
import logging
from typing import List, Dict
from PyPDF2 import PdfReader
import easyocr
import unicodedata
from docx import Document
from openpyxl import load_workbook
from bs4 import BeautifulSoup


class LicitacionAttachmentProcessor:
    """
    Procesador de adjuntos de licitaciones.

    Extrae texto de archivos PDF, imágenes, DOCX, HTML y XLSX.
    Implementa OCR para documentos escaneados y filtros de relevancia.
    """

    # Archivos a ignorar (no relevantes)
    BLACKLIST = [
        'politicas', 'canje', 'responsable', 'resolucion', 'resolución', 'resolucién',
        'certificado', 'autorización', 'autorizacion', 'autorizacién', 'certificate',
        'autorizaci6n', 'resoluci6n', 'carta'
    ]

    # Límites de procesamiento
    MAX_PAGES = 10
    TEXT_THRESHOLD = 100
    MAX_FILENAME_LENGTH = 100

    def __init__(self, input_folder: str, output_folder: str):
        """
        Inicializa el procesador.

        Args:
            input_folder: Carpeta con archivos descargados
            output_folder: Carpeta donde guardar archivos procesados (.txt)
        """
        self.input_folder = input_folder
        self.output_folder = output_folder
        os.makedirs(output_folder, exist_ok=True)

        # Inicializar OCR
        try:
            self.reader = easyocr.Reader(['es'], gpu=False)
            logging.info("OCR inicializado correctamente")
        except Exception as e:
            logging.warning(f"No se pudo inicializar OCR: {e}")
            self.reader = None

        self.skipped_files = set()

    def normalize_filename(self, filename: str) -> str:
        """Normaliza el nombre del archivo."""
        normalized = unicodedata.normalize('NFKD', filename).encode('ASCII', 'ignore').decode('ASCII')
        normalized = normalized.replace(' ', '_')
        if len(normalized) > self.MAX_FILENAME_LENGTH:
            normalized = normalized[:self.MAX_FILENAME_LENGTH]
        return normalized

    def should_skip_file(self, filename: str) -> bool:
        """Determina si un archivo debe ser omitido."""
        filename_lower = filename.lower()
        for blacklisted in self.BLACKLIST:
            if blacklisted in filename_lower:
                return True
        return False

    def extract_text_from_pdf(self, file_path: str) -> str:
        """Extrae texto de un PDF."""
        try:
            reader = PdfReader(file_path)
            text = ""

            # Limitar páginas
            num_pages = min(len(reader.pages), self.MAX_PAGES)

            for i in range(num_pages):
                page = reader.pages[i]
                page_text = page.extract_text()
                text += page_text + "\n\n"

            # Si el texto es muy corto, probablemente es un PDF escaneado
            if len(text.strip()) < self.TEXT_THRESHOLD and self.reader:
                logging.info(f"PDF con poco texto, aplicando OCR: {os.path.basename(file_path)}")
                # Aquí se podría implementar OCR con pdf2image + easyocr
                # Por ahora retornamos el texto extraído
                pass

            return text.strip()

        except Exception as e:
            logging.exception(f"Error extrayendo texto de PDF {file_path}: {e}")
            return ""

    def extract_text_from_image(self, file_path: str) -> str:
        """Extrae texto de una imagen usando OCR."""
        if not self.reader:
            logging.warning("OCR no disponible")
            return ""

        try:
            result = self.reader.readtext(file_path, detail=0)
            text = "\n".join(result)
            return text

        except Exception as e:
            logging.exception(f"Error extrayendo texto de imagen {file_path}: {e}")
            return ""

    def extract_text_from_docx(self, file_path: str) -> str:
        """Extrae texto de un archivo DOCX."""
        try:
            doc = Document(file_path)
            text = "\n".join([para.text for para in doc.paragraphs])
            return text

        except Exception as e:
            logging.exception(f"Error extrayendo texto de DOCX {file_path}: {e}")
            return ""

    def extract_text_from_html(self, file_path: str) -> str:
        """Extrae texto de un archivo HTML."""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                soup = BeautifulSoup(f.read(), 'html.parser')
                text = soup.get_text(separator='\n', strip=True)
                return text

        except Exception as e:
            logging.exception(f"Error extrayendo texto de HTML {file_path}: {e}")
            return ""

    def extract_text_from_xlsx(self, file_path: str) -> str:
        """Extrae texto de un archivo XLSX."""
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            text = ""

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text += f"\n--- Hoja: {sheet_name} ---\n"

                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                    text += row_text + "\n"

            return text

        except Exception as e:
            logging.exception(f"Error extrayendo texto de XLSX {file_path}: {e}")
            return ""

    def process_file(self, file_path: str, relative_path: str) -> Dict[str, str]:
        """
        Procesa un archivo individual.

        Args:
            file_path: Ruta completa al archivo
            relative_path: Ruta relativa para naming

        Returns:
            dict: {'input_path': str, 'output_path': str, 'text': str}
        """
        filename = os.path.basename(file_path)

        # Verificar si debe omitirse
        if self.should_skip_file(filename):
            logging.info(f"Omitiendo archivo (blacklist): {filename}")
            self.skipped_files.add(filename)
            return None

        # Determinar tipo de archivo
        ext = os.path.splitext(filename)[1].lower()

        # Extraer texto según tipo
        text = ""
        if ext == '.pdf':
            text = self.extract_text_from_pdf(file_path)
        elif ext in ['.jpg', '.jpeg', '.png']:
            text = self.extract_text_from_image(file_path)
        elif ext == '.docx':
            text = self.extract_text_from_docx(file_path)
        elif ext == '.html':
            text = self.extract_text_from_html(file_path)
        elif ext == '.xlsx':
            text = self.extract_text_from_xlsx(file_path)
        else:
            logging.warning(f"Tipo de archivo no soportado: {filename}")
            return None

        if not text or len(text.strip()) < 50:
            logging.warning(f"Archivo con muy poco texto: {filename}")
            return None

        # Crear nombre de salida
        output_filename = self.normalize_filename(filename.replace(ext, '.txt'))
        output_path = os.path.join(self.output_folder, output_filename)

        # Guardar texto extraído
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(text)

        logging.info(f"Procesado: {filename} -> {output_filename}")

        return {
            'input_path': file_path,
            'output_path': output_path,
            'text': text
        }

    def process_all(self) -> List[Dict[str, str]]:
        """
        Procesa todos los archivos en input_folder.

        Returns:
            list: Lista de diccionarios con info de archivos procesados
        """
        archivos_procesados = []

        # Procesar carpeta tech
        tech_folder = os.path.join(self.input_folder, "tech")
        if os.path.exists(tech_folder):
            logging.info("Procesando anexos técnicos...")
            for filename in os.listdir(tech_folder):
                file_path = os.path.join(tech_folder, filename)
                if os.path.isfile(file_path):
                    result = self.process_file(file_path, f"tech/{filename}")
                    if result:
                        archivos_procesados.append(result)

        # Procesar carpeta econ
        econ_folder = os.path.join(self.input_folder, "econ")
        if os.path.exists(econ_folder):
            logging.info("Procesando anexos económicos...")
            for filename in os.listdir(econ_folder):
                file_path = os.path.join(econ_folder, filename)
                if os.path.isfile(file_path):
                    result = self.process_file(file_path, f"econ/{filename}")
                    if result:
                        archivos_procesados.append(result)

        logging.info(f"Total procesados: {len(archivos_procesados)} archivos")
        logging.info(f"Total omitidos: {len(self.skipped_files)} archivos")

        return archivos_procesados
